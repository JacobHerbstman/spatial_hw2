#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
import re
import sys
import zipfile
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Iterable
from urllib.parse import urljoin

import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
INPUT_DIR = ROOT / "input"
INTERMEDIATE_DIR = ROOT / "intermediate"
TIMEOUT = 90
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0 Safari/537.36"
    )
}

ONET_URL = "https://www.onetcenter.org/dl_files/database/db_29_1_text.zip"
OEWS_URLS = [
    "https://www.bls.gov/oes/special-requests/oesm23nat.zip",
    "https://www.bls.gov/oes/special-requests/oes_research_2023.xlsx",
]
OEWS_RESEARCH_PAGES = [
    "https://www.bls.gov/oes/oes_research_2023.htm",
    "https://www.bls.gov/oes/current/oes_research_2023.htm",
    "https://www.bls.gov/oes/current/oessrci.htm",
]
QCEW_URL = "https://data.bls.gov/cew/data/files/2023/csv/2023_annual_by_area.zip"
ONET_MEMBERS = {"Work Activities.txt", "Occupation Data.txt"}
MANUAL_OEWS_NAMES = [
    "oesm23nat.zip",
    "oes_research_2023.xlsx",
    "nat4d_M2023_dl.xlsx",
]


@dataclass
class ProvenanceRow:
    dataset: str
    url: str
    status: str
    http_status: int | str
    local_path: str
    message: str


def log(message: str) -> None:
    print(message, flush=True)


def ensure_dirs() -> None:
    for path in (INPUT_DIR, INTERMEDIATE_DIR):
        path.mkdir(parents=True, exist_ok=True)


def write_provenance(rows: list[ProvenanceRow]) -> None:
    json_path = INTERMEDIATE_DIR / "download_provenance.json"
    csv_path = INTERMEDIATE_DIR / "download_provenance.csv"
    payload = [asdict(row) for row in rows]
    json_path.write_text(json.dumps(payload, indent=2))
    with csv_path.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(payload[0].keys()) if payload else list(ProvenanceRow.__annotations__.keys()))
        writer.writeheader()
        writer.writerows(payload)


def session_get(url: str, stream: bool = False) -> requests.Response:
    return requests.get(url, headers=HEADERS, timeout=TIMEOUT, allow_redirects=True, stream=stream)


def download_to_path(url: str, dest: Path) -> tuple[Path | None, int | str, str]:
    try:
        response = session_get(url, stream=True)
    except requests.RequestException as exc:
        return None, "request_error", str(exc)

    if response.status_code != 200:
        return None, response.status_code, f"Download failed for {url} with HTTP {response.status_code}"

    dest.parent.mkdir(parents=True, exist_ok=True)
    with dest.open("wb") as handle:
        for chunk in response.iter_content(chunk_size=1 << 20):
            if chunk:
                handle.write(chunk)
    return dest, response.status_code, "downloaded"


def extract_zip_members(zip_path: Path, members: Iterable[str], dest_dir: Path) -> list[Path]:
    extracted: list[Path] = []
    with zipfile.ZipFile(zip_path) as archive:
        lower_lookup = {Path(name).name.lower(): name for name in archive.namelist()}
        for member in members:
            archive_name = lower_lookup.get(member.lower())
            if archive_name is None:
                raise FileNotFoundError(f"Could not find {member} in {zip_path.name}")
            dest_path = dest_dir / Path(archive_name).name
            if not dest_path.exists():
                with archive.open(archive_name) as src, dest_path.open("wb") as dst:
                    dst.write(src.read())
            extracted.append(dest_path)
    return extracted


def extract_first_matching(zip_path: Path, suffix: str, dest_dir: Path, name_pattern: str | None = None) -> Path:
    with zipfile.ZipFile(zip_path) as archive:
        members = [name for name in archive.namelist() if name.lower().endswith(suffix.lower())]
        if name_pattern is not None:
            members = [name for name in members if re.search(name_pattern, Path(name).name, flags=re.IGNORECASE)]
        if not members:
            raise FileNotFoundError(f"No {suffix} member found in {zip_path.name}")
        member = sorted(members)[0]
        dest_path = dest_dir / Path(member).name
        if not dest_path.exists():
            with archive.open(member) as src, dest_path.open("wb") as dst:
                dst.write(src.read())
    return dest_path


def workbook_has_industry_detail(path: Path) -> bool:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return False
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row is None:
            continue
        header_lookup = {
            normalize_header(value): idx for idx, value in enumerate(header_row) if normalize_header(value)
        }
        naics_idx = header_lookup.get("NAICS")
        if naics_idx is None:
            continue
        for row in sheet.iter_rows(min_row=2, max_row=500, values_only=True):
            if row is None or naics_idx >= len(row):
                continue
            digits = canonicalize_code(row[naics_idx])
            if digits and digits not in {"0", "00", "000", "0000", "00000", "000000"}:
                return True
    return False


def canonicalize_code(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return None
    if re.fullmatch(r"\d+(\.0+)?", text):
        return text.split(".")[0]
    return None


def normalize_header(value: object) -> str:
    text = "" if value is None else str(value).strip().upper()
    return re.sub(r"[^A-Z0-9]+", "_", text).strip("_")


def find_existing_oews() -> Path | None:
    patterns = [
        "oes_research_2023.xlsx",
        "nat4d_M2023_dl.xlsx",
        "*2023*.xlsx",
    ]
    for pattern in patterns:
        matches = sorted(INPUT_DIR.rglob(pattern))
        if matches:
            for match in matches:
                if workbook_has_industry_detail(match):
                    return match
    for zip_name in ("oesm23nat.zip",):
        zip_path = INPUT_DIR / zip_name
        if zip_path.exists():
            try:
                workbook_path = extract_first_matching(zip_path, ".xlsx", INPUT_DIR, name_pattern=r"2023|M2023")
                if workbook_has_industry_detail(workbook_path):
                    return workbook_path
            except Exception:
                continue
    return None


def discover_scraped_oews_urls() -> list[str]:
    discovered: list[str] = []
    for page_url in OEWS_RESEARCH_PAGES:
        try:
            response = session_get(page_url)
        except requests.RequestException:
            continue
        if response.status_code != 200:
            continue
        hrefs = re.findall(r'href=["\']([^"\']+)["\']', response.text, flags=re.IGNORECASE)
        for href in hrefs:
            if not re.search(r"2023.*\.(xlsx|zip)$", href, flags=re.IGNORECASE):
                continue
            if "research" not in href.lower() and "nat" not in href.lower():
                continue
            discovered.append(urljoin(page_url, href))
    deduped: list[str] = []
    seen: set[str] = set()
    for url in discovered:
        if url not in seen:
            seen.add(url)
            deduped.append(url)
    return deduped


def write_oews_manual_instructions() -> Path:
    path = INTERMEDIATE_DIR / "oews_manual_download_instructions.txt"
    lines = [
        "Unable to retrieve a strict-2023 OEWS occupation-by-industry workbook automatically.",
        "",
        "Please manually download one of the following into the task input directory:",
        f"  {INPUT_DIR}",
        "",
        "Accepted filenames:",
    ]
    lines.extend(f"  - {name}" for name in MANUAL_OEWS_NAMES)
    lines.extend(
        [
            "",
            "Required content: a 2023 OEWS/OES national occupation-by-industry workbook",
            "with columns including OCC_CODE, NAICS, and TOT_EMP.",
            "",
            "Attempted direct URLs:",
            f"  - {OEWS_URLS[0]}",
            f"  - {OEWS_URLS[1]}",
        ]
    )
    path.write_text("\n".join(lines) + "\n")
    return path


def ensure_onet(rows: list[ProvenanceRow]) -> None:
    zip_path = INPUT_DIR / "db_29_1_text.zip"
    extracted_paths = [INPUT_DIR / name for name in ONET_MEMBERS]
    if zip_path.exists() and all(path.exists() for path in extracted_paths):
        rows.append(ProvenanceRow("onet", ONET_URL, "existing", "", str(zip_path), "Using existing zip and extracted text files"))
        return
    if zip_path.exists():
        extracted = extract_zip_members(zip_path, ONET_MEMBERS, INPUT_DIR)
        rows.append(ProvenanceRow("onet", ONET_URL, "existing", "", str(zip_path), "Using existing O*NET zip"))
        rows.append(
            ProvenanceRow(
                "onet_extract",
                ONET_URL,
                "extracted",
                "",
                str(extracted[0].parent),
                f"Extracted {', '.join(path.name for path in extracted)} from existing zip",
            )
        )
        return

    log(f"Downloading O*NET data from {ONET_URL}")
    downloaded_path, http_status, message = download_to_path(ONET_URL, zip_path)
    rows.append(ProvenanceRow("onet", ONET_URL, "downloaded" if downloaded_path else "failed", http_status, str(zip_path), message))
    if downloaded_path is None:
        raise RuntimeError(message)
    extracted = extract_zip_members(zip_path, ONET_MEMBERS, INPUT_DIR)
    rows.append(
        ProvenanceRow(
            "onet_extract",
            ONET_URL,
            "extracted",
            "",
            str(extracted[0].parent),
            f"Extracted {', '.join(path.name for path in extracted)}",
        )
    )


def ensure_qcew(rows: list[ProvenanceRow]) -> None:
    zip_path = INPUT_DIR / "2023_annual_by_area.zip"
    if zip_path.exists():
        rows.append(ProvenanceRow("qcew", QCEW_URL, "existing", "", str(zip_path), "Using existing QCEW by-area zip"))
        return

    log(f"Downloading QCEW annual by-area data from {QCEW_URL}")
    downloaded_path, http_status, message = download_to_path(QCEW_URL, zip_path)
    rows.append(ProvenanceRow("qcew", QCEW_URL, "downloaded" if downloaded_path else "failed", http_status, str(zip_path), message))
    if downloaded_path is None:
        raise RuntimeError(message)


def ensure_oews(rows: list[ProvenanceRow]) -> None:
    existing = find_existing_oews()
    if existing is not None:
        rows.append(ProvenanceRow("oews", "", "existing", "", str(existing), "Using existing 2023 OEWS workbook"))
        return

    attempted_urls = list(OEWS_URLS)
    attempted_urls.extend(discover_scraped_oews_urls())

    seen: set[str] = set()
    for url in attempted_urls:
        if url in seen:
            continue
        seen.add(url)
        filename = Path(url.split("?")[0]).name or "oews_2023_download"
        dest = INPUT_DIR / filename
        log(f"Attempting strict-2023 OEWS download from {url}")
        downloaded_path, http_status, message = download_to_path(url, dest)
        rows.append(
            ProvenanceRow(
                "oews",
                url,
                "downloaded" if downloaded_path else "failed",
                http_status,
                str(dest),
                message,
            )
        )
        if downloaded_path is None:
            continue

        if downloaded_path.suffix.lower() == ".zip":
            try:
                workbook_path = extract_first_matching(downloaded_path, ".xlsx", INPUT_DIR, name_pattern=r"2023|M2023")
            except Exception as exc:
                rows.append(ProvenanceRow("oews_extract", url, "failed", "", str(downloaded_path), str(exc)))
                continue
            if not workbook_has_industry_detail(workbook_path):
                rows.append(
                    ProvenanceRow(
                        "oews_validate",
                        url,
                        "failed",
                        "",
                        str(workbook_path),
                        "Workbook does not contain industry detail beyond NAICS 000000",
                    )
                )
                continue
            rows.append(
                ProvenanceRow(
                    "oews_extract",
                    url,
                    "extracted",
                    "",
                    str(workbook_path),
                    f"Extracted {workbook_path.name}",
                )
            )
            return

        if downloaded_path.suffix.lower() == ".xlsx":
            if not workbook_has_industry_detail(downloaded_path):
                rows.append(
                    ProvenanceRow(
                        "oews_validate",
                        url,
                        "failed",
                        "",
                        str(downloaded_path),
                        "Workbook does not contain industry detail beyond NAICS 000000",
                    )
                )
                continue
            return

    instruction_path = write_oews_manual_instructions()
    log(
        "Strict-2023 OEWS download failed. "
        f"Manual instructions written to {instruction_path}"
    )
    raise RuntimeError(
        "Unable to download strict-2023 OEWS data automatically. "
        f"See {instruction_path}"
    )


def main() -> None:
    ensure_dirs()
    rows: list[ProvenanceRow] = []
    try:
        ensure_onet(rows)
        ensure_oews(rows)
        ensure_qcew(rows)
    except Exception as exc:
        rows.append(ProvenanceRow("summary", "", "failed", "", "", str(exc)))
        write_provenance(rows)
        log(str(exc))
        sys.exit(1)

    rows.append(ProvenanceRow("summary", "", "success", "", "", "All required downloads are available"))
    write_provenance(rows)
    log(f"Wrote download provenance to {INTERMEDIATE_DIR / 'download_provenance.json'}")


if __name__ == "__main__":
    main()
