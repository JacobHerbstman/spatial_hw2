#!/usr/bin/env python3
from __future__ import annotations

import json
import re
import sys
import zipfile
from datetime import datetime, timezone
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
INPUT_DIR = ROOT / "input"
INTERMEDIATE_DIR = ROOT / "intermediate"
OUTPUT_DIR = ROOT / "output"
METADATA_PATH = OUTPUT_DIR / "cognitive_intensity_metadata.json"

ONET_ACTIVITY_PATH = INPUT_DIR / "Work Activities.txt"
ONET_OCC_PATH = INPUT_DIR / "Occupation Data.txt"
PROVENANCE_PATH = INTERMEDIATE_DIR / "download_provenance.json"

CNR_ELEMENT_IDS = {
    "4.A.2.a.4": "Analyzing Data or Information",
    "4.A.2.b.2": "Thinking Creatively",
    "4.A.4.a.1": "Interpreting the Meaning of Information for Others",
    "4.A.4.a.4": "Establishing and Maintaining Interpersonal Relationships",
    "4.A.4.b.4": "Guiding, Directing, and Motivating Subordinates",
}
MATCH_SHARE_THRESHOLD = 0.85

FIPS_TO_STATE = {
    "01": ("AL", 1),
    "02": ("AK", 2),
    "04": ("AZ", 3),
    "05": ("AR", 4),
    "06": ("CA", 5),
    "08": ("CO", 6),
    "09": ("CT", 7),
    "10": ("DE", 8),
    "12": ("FL", 9),
    "13": ("GA", 10),
    "15": ("HI", 11),
    "16": ("ID", 12),
    "17": ("IL", 13),
    "18": ("IN", 14),
    "19": ("IA", 15),
    "20": ("KS", 16),
    "21": ("KY", 17),
    "22": ("LA", 18),
    "23": ("ME", 19),
    "24": ("MD", 20),
    "25": ("MA", 21),
    "26": ("MI", 22),
    "27": ("MN", 23),
    "28": ("MS", 24),
    "29": ("MO", 25),
    "30": ("MT", 26),
    "31": ("NE", 27),
    "32": ("NV", 28),
    "33": ("NH", 29),
    "34": ("NJ", 30),
    "35": ("NM", 31),
    "36": ("NY", 32),
    "37": ("NC", 33),
    "38": ("ND", 34),
    "39": ("OH", 35),
    "40": ("OK", 36),
    "41": ("OR", 37),
    "42": ("PA", 38),
    "44": ("RI", 39),
    "45": ("SC", 40),
    "46": ("SD", 41),
    "47": ("TN", 42),
    "48": ("TX", 43),
    "49": ("UT", 44),
    "50": ("VT", 45),
    "51": ("VA", 46),
    "53": ("WA", 47),
    "54": ("WV", 48),
    "55": ("WI", 49),
    "56": ("WY", 50),
}
STATE_ABBR_TO_IDX = {abbr: idx for _, (abbr, idx) in FIPS_TO_STATE.items()}
STATE_NAME_TO_ABBR = {
    "ALABAMA": "AL",
    "ALASKA": "AK",
    "ARIZONA": "AZ",
    "ARKANSAS": "AR",
    "CALIFORNIA": "CA",
    "COLORADO": "CO",
    "CONNECTICUT": "CT",
    "DELAWARE": "DE",
    "DISTRICT OF COLUMBIA": None,
    "FLORIDA": "FL",
    "GEORGIA": "GA",
    "HAWAII": "HI",
    "IDAHO": "ID",
    "ILLINOIS": "IL",
    "INDIANA": "IN",
    "IOWA": "IA",
    "KANSAS": "KS",
    "KENTUCKY": "KY",
    "LOUISIANA": "LA",
    "MAINE": "ME",
    "MARYLAND": "MD",
    "MASSACHUSETTS": "MA",
    "MICHIGAN": "MI",
    "MINNESOTA": "MN",
    "MISSISSIPPI": "MS",
    "MISSOURI": "MO",
    "MONTANA": "MT",
    "NEBRASKA": "NE",
    "NEVADA": "NV",
    "NEW HAMPSHIRE": "NH",
    "NEW JERSEY": "NJ",
    "NEW MEXICO": "NM",
    "NEW YORK": "NY",
    "NORTH CAROLINA": "NC",
    "NORTH DAKOTA": "ND",
    "OHIO": "OH",
    "OKLAHOMA": "OK",
    "OREGON": "OR",
    "PENNSYLVANIA": "PA",
    "RHODE ISLAND": "RI",
    "SOUTH CAROLINA": "SC",
    "SOUTH DAKOTA": "SD",
    "TENNESSEE": "TN",
    "TEXAS": "TX",
    "UTAH": "UT",
    "VERMONT": "VT",
    "VIRGINIA": "VA",
    "WASHINGTON": "WA",
    "WEST VIRGINIA": "WV",
    "WISCONSIN": "WI",
    "WYOMING": "WY",
}

CDP_SECTOR_MAP = {
    1: {"name": "Manufacturing", "naics_prefixes": {"31", "32", "33"}},
    2: {"name": "Construction", "naics_prefixes": {"23"}},
    3: {"name": "Wholesale/Retail", "naics_prefixes": {"42", "44", "45"}},
    4: {
        "name": "Services",
        "naics_prefixes": {"48", "49", "51", "52", "53", "54", "55", "56", "61", "62", "71", "72", "81"},
    },
}
QCEW_BY_AREA_AGGLVL_BY_SIGLEN = {3: "55", 4: "56", 5: "57", 6: "58"}
QCEW_DETAILED_OWN_CODES = {"1", "2", "3", "5"}
OEWS_COLUMN_ALIASES = {
    "OCC_CODE": {"OCC_CODE", "OCC CODE", "SOC_CODE", "SOC CODE"},
    "OCC_TITLE": {"OCC_TITLE", "OCC TITLE", "SOC_TITLE", "SOC TITLE"},
    "NAICS": {"NAICS", "INDUSTRY_CODE", "INDUSTRY CODE"},
    "NAICS_TITLE": {"NAICS_TITLE", "NAICS TITLE", "INDUSTRY_TITLE", "INDUSTRY TITLE"},
    "TOT_EMP": {"TOT_EMP", "TOT EMP", "EMPLOYMENT"},
    "O_GROUP": {"O_GROUP", "O GROUP", "OCC_GROUP", "OCC GROUP"},
}
STATE_OEWS_COLUMN_ALIASES = {
    **OEWS_COLUMN_ALIASES,
    "AREA": {"AREA", "AREA_CODE", "AREA CODE", "STATE_CODE", "STATE CODE"},
    "AREA_TITLE": {"AREA_TITLE", "AREA TITLE", "AREA_NAME", "AREA NAME", "STATE", "STATE_NAME", "STATE NAME"},
    "PRIM_STATE": {"PRIM_STATE", "PRIM STATE", "ST", "STATE_ABBR", "STATE ABBR"},
}


def log(message: str) -> None:
    print(message, flush=True)


def normalize_header(value: object) -> str:
    text = "" if value is None else str(value).strip().upper()
    return re.sub(r"[^A-Z0-9]+", "_", text).strip("_")


def canonicalize_code(value: object) -> str | None:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return None
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return None
    if re.fullmatch(r"\d+(\.0+)?", text):
        return text.split(".")[0]
    return None


def to_numeric(series: pd.Series) -> pd.Series:
    cleaned = (
        series.astype(str)
        .str.replace(",", "", regex=False)
        .str.replace("$", "", regex=False)
        .str.replace(" ", "", regex=False)
    )
    return pd.to_numeric(cleaned, errors="coerce")


def sig_code(digits: str) -> str:
    stripped = digits.rstrip("0")
    return stripped if stripped else digits


def map_sector_from_naics(code: str) -> tuple[int, str] | None:
    prefix = code[:2]
    for sector_idx, meta in CDP_SECTOR_MAP.items():
        if prefix in meta["naics_prefixes"]:
            return sector_idx, meta["name"]
    return None


def weighted_mean(values: pd.Series, weights: pd.Series) -> float:
    mask = values.notna() & weights.notna() & (weights > 0)
    if not mask.any():
        return float("nan")
    return float(np.average(values[mask], weights=weights[mask]))


def longest_available_prefix(code: str, available_codes: set[str]) -> str | None:
    for length in range(len(code), 2, -1):
        candidate = code[:length]
        if candidate in available_codes:
            return candidate
    return None


def locate_oews_workbook() -> Path:
    patterns = ["*2023*.xlsx", "nat4d_M2023_dl.xlsx", "oes_research_2023.xlsx"]
    for pattern in patterns:
        matches = sorted(INPUT_DIR.rglob(pattern))
        for match in matches:
            if workbook_has_industry_detail(match):
                return match
    zip_matches = sorted(INPUT_DIR.rglob("oesm23nat.zip"))
    for zip_path in zip_matches:
        with zipfile.ZipFile(zip_path) as archive:
            workbook_members = [name for name in archive.namelist() if name.lower().endswith(".xlsx")]
            if workbook_members:
                member = sorted(workbook_members)[0]
                dest = INPUT_DIR / Path(member).name
                if not dest.exists():
                    with archive.open(member) as src, dest.open("wb") as dst:
                        dst.write(src.read())
                if workbook_has_industry_detail(dest):
                    return dest
    raise FileNotFoundError("Could not locate a strict-2023 OEWS workbook in input/")


def load_download_provenance() -> list[dict]:
    if not PROVENANCE_PATH.exists():
        return []
    return json.loads(PROVENANCE_PATH.read_text())


def collapse_onet_scores() -> pd.DataFrame:
    if not ONET_ACTIVITY_PATH.exists() or not ONET_OCC_PATH.exists():
        raise FileNotFoundError("Missing O*NET text files. Run the download step first.")

    activities = pd.read_csv(ONET_ACTIVITY_PATH, sep="\t", dtype=str, low_memory=False)
    occupations = pd.read_csv(ONET_OCC_PATH, sep="\t", dtype=str, low_memory=False)
    activities = activities[
        (activities["Scale ID"] == "IM")
        & (activities["Element ID"].isin(CNR_ELEMENT_IDS))
    ].copy()
    activities["data_value"] = pd.to_numeric(activities["Data Value"], errors="coerce")
    activities["soc_code"] = activities["O*NET-SOC Code"].str.replace(r"\.\d+$", "", regex=True)

    occ_scores = (
        activities.groupby(["soc_code"], as_index=False)
        .agg(
            cognitive_score_raw=("data_value", "mean"),
            item_count=("Element ID", "nunique"),
        )
        .sort_values("soc_code")
    )
    occ_scores = occ_scores[occ_scores["item_count"] > 0].copy()
    min_score = occ_scores["cognitive_score_raw"].min()
    max_score = occ_scores["cognitive_score_raw"].max()
    if pd.isna(min_score) or pd.isna(max_score) or max_score <= min_score:
        raise RuntimeError("O*NET cognitive score normalization failed because scores are degenerate.")
    occ_scores["cognitive_score"] = (occ_scores["cognitive_score_raw"] - min_score) / (max_score - min_score)

    occupation_titles = (
        occupations.assign(soc_code=occupations["O*NET-SOC Code"].str.replace(r"\.\d+$", "", regex=True))
        .groupby("soc_code", as_index=False)["Title"]
        .first()
        .rename(columns={"Title": "occupation_title"})
    )
    occ_scores = occ_scores.merge(occupation_titles, on="soc_code", how="left")
    occ_scores = occ_scores[["soc_code", "occupation_title", "item_count", "cognitive_score_raw", "cognitive_score"]]
    occ_scores.to_csv(INTERMEDIATE_DIR / "cognitive_score_by_soc.csv", index=False)
    log(f"Wrote occupation cognitive scores to {INTERMEDIATE_DIR / 'cognitive_score_by_soc.csv'}")
    return occ_scores


def read_excel_table(path: Path, alias_map: dict[str, set[str]]) -> pd.DataFrame:
    workbook = load_workbook(path, read_only=True, data_only=True)
    best_sheet_name: str | None = None
    best_header_row: int | None = None
    best_positions: dict[str, int] | None = None
    best_score = -1

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=25, values_only=True), start=1):
            positions = header_positions(row, alias_map)
            if not required_positions_present(positions):
                continue
            score = sample_score_sheet(sheet, row_number, positions)
            if score > best_score:
                best_score = score
                best_sheet_name = sheet_name
                best_header_row = row_number
                best_positions = positions
            break

    if best_sheet_name is None or best_header_row is None or best_positions is None:
        raise RuntimeError(f"Could not find a usable data table in {path.name}")

    records: list[dict[str, object]] = []
    sheet = workbook[best_sheet_name]
    for row in sheet.iter_rows(min_row=best_header_row + 1, values_only=True):
        if row is None:
            continue
        record = {
            canonical: row[idx] if idx < len(row) else None
            for canonical, idx in best_positions.items()
        }
        records.append(record)
    return pd.DataFrame.from_records(records)


def header_positions(row: tuple[object, ...], alias_map: dict[str, set[str]]) -> dict[str, int]:
    positions: dict[str, int] = {}
    for idx, value in enumerate(row):
        normalized = normalize_header(value)
        if not normalized:
            continue
        for canonical, aliases in alias_map.items():
            if normalized in aliases and canonical not in positions:
                positions[canonical] = idx
                break
    return positions


def required_positions_present(positions: dict[str, int]) -> bool:
    return all(key in positions for key in ("OCC_CODE", "NAICS", "TOT_EMP"))


def sample_score_sheet(sheet, header_row: int, positions: dict[str, int]) -> int:
    occ_idx = positions["OCC_CODE"]
    naics_idx = positions["NAICS"]
    score = 0
    for row in sheet.iter_rows(min_row=header_row + 1, max_row=header_row + 500, values_only=True):
        if row is None:
            continue
        occ_value = row[occ_idx] if occ_idx < len(row) else None
        naics_value = row[naics_idx] if naics_idx < len(row) else None
        occ_code = "" if occ_value is None else str(occ_value).strip()
        if re.fullmatch(r"\d{2}-\d{4}", occ_code) and canonicalize_code(naics_value) is not None:
            score += 1
    return score


def workbook_has_industry_detail(path: Path) -> bool:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return False
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row is None:
            continue
        header_lookup = {
            normalize_header(value): idx for idx, value in enumerate(header_row) if normalize_header(value)
        }
        naics_idx = header_lookup.get("NAICS")
        if naics_idx is None:
            continue
        for row in sheet.iter_rows(min_row=2, max_row=500, values_only=True):
            if row is None or naics_idx >= len(row):
                continue
            digits = canonicalize_code(row[naics_idx])
            if digits and digits not in {"0", "00", "000", "0000", "00000", "000000"}:
                return True
    return False


def rename_columns(df: pd.DataFrame, alias_map: dict[str, set[str]]) -> pd.DataFrame:
    renamed: dict[str, str] = {}
    for column in df.columns:
        normalized = normalize_header(column)
        for canonical, aliases in alias_map.items():
            if normalized in aliases:
                renamed[column] = canonical
                break
    return df.rename(columns=renamed)


def score_candidate_table(df: pd.DataFrame) -> int:
    if "OCC_CODE" not in df.columns or "NAICS" not in df.columns or "TOT_EMP" not in df.columns:
        return -1
    occ = df["OCC_CODE"].astype(str).str.strip().str.fullmatch(r"\d{2}-\d{4}")
    naics = df["NAICS"].map(canonicalize_code).notna()
    return int((occ & naics).sum())


def build_leaf_codes(raw_codes: pd.Series) -> set[str]:
    codes = sorted({sig_code(code) for code in raw_codes if code})
    leaves: set[str] = set()
    for code in codes:
        has_child = any(other.startswith(code) and len(other) > len(code) for other in codes)
        if not has_child:
            leaves.add(code)
    return leaves


def load_oews_national_matrix() -> pd.DataFrame:
    workbook_path = locate_oews_workbook()
    log(f"Reading OEWS national workbook {workbook_path}")
    raw = read_excel_table(workbook_path, OEWS_COLUMN_ALIASES)
    missing = {"OCC_CODE", "NAICS", "TOT_EMP"} - set(raw.columns)
    if missing:
        raise RuntimeError(f"OEWS workbook is missing required columns: {sorted(missing)}")

    df = raw.copy()
    if "O_GROUP" in df.columns:
        df["o_group"] = df["O_GROUP"].astype(str).str.strip().str.lower()
        df = df[df["o_group"] == "detailed"].copy()
    df["soc_code"] = df["OCC_CODE"].astype(str).str.strip()
    df = df[df["soc_code"].str.fullmatch(r"\d{2}-\d{4}")].copy()
    df["naics_digits"] = df["NAICS"].map(canonicalize_code)
    df = df[df["naics_digits"].notna()].copy()
    if not any(code not in {"0", "00", "000", "0000", "00000", "000000"} for code in df["naics_digits"]):
        raise RuntimeError(
            "The detected OEWS workbook only contains cross-industry rows (NAICS 000000). "
            "Place a 2023 occupation-by-industry workbook with industry detail in input/."
        )
    df["naics_sig"] = df["naics_digits"].map(sig_code)
    df["sig_len"] = df["naics_sig"].str.len()
    df = df[df["sig_len"].between(3, 6)].copy()
    leaf_codes = build_leaf_codes(df["naics_digits"])
    df = df[df["naics_sig"].isin(leaf_codes)].copy()
    df["tot_emp"] = to_numeric(df["TOT_EMP"])
    df = df[df["tot_emp"] > 0].copy()
    if "NAICS_TITLE" in df.columns:
        df["naics_title"] = df["NAICS_TITLE"].astype(str).str.strip()
    else:
        df["naics_title"] = df["naics_sig"]
    if "OCC_TITLE" in df.columns:
        df["occupation_title"] = df["OCC_TITLE"].astype(str).str.strip()
    else:
        df["occupation_title"] = df["soc_code"]
    sector_meta = df["naics_sig"].map(map_sector_from_naics)
    df = df[sector_meta.notna()].copy()
    df["sector_idx"] = [meta[0] for meta in sector_meta[sector_meta.notna()]]
    df["sector_name"] = [meta[1] for meta in sector_meta[sector_meta.notna()]]
    return df[
        [
            "soc_code",
            "occupation_title",
            "naics_sig",
            "naics_title",
            "sector_idx",
            "sector_name",
            "tot_emp",
        ]
    ].copy()


def merge_oews_with_cognitive_scores(oews: pd.DataFrame, occ_scores: pd.DataFrame) -> tuple[pd.DataFrame, float]:
    merged = oews.merge(occ_scores, on="soc_code", how="left")
    matched_emp = merged.loc[merged["cognitive_score"].notna(), "tot_emp"].sum()
    total_emp = merged["tot_emp"].sum()
    match_share = float(matched_emp / total_emp) if total_emp > 0 else float("nan")
    log(f"O*NET to OEWS employment match share: {match_share:.4f}")
    if not np.isfinite(match_share) or match_share < MATCH_SHARE_THRESHOLD:
        raise RuntimeError(
            f"O*NET/OEWS employment match share {match_share:.4f} is below the threshold "
            f"{MATCH_SHARE_THRESHOLD:.2f}"
        )
    merged = merged[merged["cognitive_score"].notna()].copy()
    return merged, match_share


def compute_national_outputs(merged: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, float]:
    industry_rows: list[dict] = []
    for (naics_sig, sector_idx, sector_name), group in merged.groupby(["naics_sig", "sector_idx", "sector_name"], sort=True):
        industry_rows.append(
            {
                "naics_sig": naics_sig,
                "naics_title": group["naics_title"].dropna().iloc[0],
                "sector_idx": sector_idx,
                "sector_name": sector_name,
                "employment": float(group["tot_emp"].sum()),
                "cognitive_intensity_raw": weighted_mean(group["cognitive_score"], group["tot_emp"]),
            }
        )
    industry_df = pd.DataFrame(industry_rows).sort_values(["sector_idx", "naics_sig"]).reset_index(drop=True)
    industry_df.to_csv(INTERMEDIATE_DIR / "national_industry_cognitive.csv", index=False)
    log(f"Wrote national industry intensities to {INTERMEDIATE_DIR / 'national_industry_cognitive.csv'}")

    sector_rows: list[dict] = []
    for (sector_idx, sector_name), group in merged.groupby(["sector_idx", "sector_name"], sort=True):
        sector_rows.append(
            {
                "sector_idx": sector_idx,
                "sector_name": sector_name,
                "employment": float(group["tot_emp"].sum()),
                "cognitive_intensity_raw": weighted_mean(group["cognitive_score"], group["tot_emp"]),
            }
        )
    sector_df = pd.DataFrame(sector_rows).sort_values("sector_idx").reset_index(drop=True)
    sector_df.to_csv(INTERMEDIATE_DIR / "national_sector_cognitive.csv", index=False)
    log(f"Wrote national sector intensities to {INTERMEDIATE_DIR / 'national_sector_cognitive.csv'}")

    national_overall = weighted_mean(merged["cognitive_score"], merged["tot_emp"])
    return industry_df, sector_df, national_overall


def locate_qcew_by_area_zip() -> Path:
    matches = sorted(INPUT_DIR.glob("*annual_by_area*.zip"))
    if not matches:
        raise FileNotFoundError("Could not find the 2023 QCEW by-area zip in input/")
    return matches[0]


def load_qcew_state_industry() -> pd.DataFrame:
    qcew_path = locate_qcew_by_area_zip()
    log(f"Reading QCEW state-by-industry data from {qcew_path}")
    usecols = ["area_fips", "own_code", "industry_code", "agglvl_code", "annual_avg_emplvl"]
    frames: list[pd.DataFrame] = []
    with zipfile.ZipFile(qcew_path) as archive:
        members = [
            name
            for name in archive.namelist()
            if name.endswith(".csv") and re.search(r"\b\d{5} .*Statewide\.csv$", name)
        ]
        for member in members:
            area_match = re.search(r"(\d{5})", Path(member).name)
            if area_match is None:
                continue
            area_code = area_match.group(1)
            if not area_code.endswith("000"):
                continue
            state_fips = area_code[:2]
            if state_fips not in FIPS_TO_STATE:
                continue
            frame = pd.read_csv(archive.open(member), usecols=usecols, dtype=str, low_memory=False)
            frame = frame[frame["own_code"].isin(QCEW_DETAILED_OWN_CODES)].copy()
            frame["industry_digits"] = frame["industry_code"].map(canonicalize_code)
            frame = frame[frame["industry_digits"].notna()].copy()
            frame["naics_sig"] = frame["industry_digits"].map(sig_code)
            frame["sig_len"] = frame["naics_sig"].str.len()
            frame = frame[frame["sig_len"].between(3, 6)].copy()
            frame["expected_agglvl"] = frame["sig_len"].map(QCEW_BY_AREA_AGGLVL_BY_SIGLEN)
            frame = frame[frame["agglvl_code"] == frame["expected_agglvl"]].copy()
            frame["employment"] = to_numeric(frame["annual_avg_emplvl"])
            frame = frame[frame["employment"] > 0].copy()
            frame["state_fips"] = state_fips
            frames.append(frame[["state_fips", "naics_sig", "employment"]])
    if not frames:
        raise RuntimeError("No usable QCEW rows were found after filtering.")
    qcew = pd.concat(frames, ignore_index=True)
    qcew = qcew.groupby(["state_fips", "naics_sig"], as_index=False)["employment"].sum()
    return qcew


def qcew_coverage_metadata(qcew: pd.DataFrame, available_codes: set[str]) -> dict:
    qcew = qcew.copy()
    sector_meta = qcew["naics_sig"].map(map_sector_from_naics)
    qcew["sector_idx"] = sector_meta.map(lambda value: value[0] if value else np.nan)
    qcew["sector_name"] = sector_meta.map(lambda value: value[1] if value else None)
    qcew["exact_match"] = qcew["naics_sig"].isin(available_codes)
    qcew["matched_naics_sig"] = qcew["naics_sig"].map(lambda code: longest_available_prefix(code, available_codes))
    qcew["prefix_match"] = qcew["matched_naics_sig"].notna()

    def summarize(frame: pd.DataFrame) -> dict:
        employment = float(frame["employment"].sum())
        exact_emp = float(frame.loc[frame["exact_match"], "employment"].sum())
        prefix_emp = float(frame.loc[frame["prefix_match"], "employment"].sum())
        return {
            "employment": employment,
            "exact_match_employment": exact_emp,
            "prefix_match_employment": prefix_emp,
            "exact_match_share": float(exact_emp / employment) if employment > 0 else float("nan"),
            "prefix_match_share": float(prefix_emp / employment) if employment > 0 else float("nan"),
        }

    cdp_mask = qcew["sector_idx"].notna()
    by_sector: list[dict] = []
    for sector_idx, meta in CDP_SECTOR_MAP.items():
        sector_frame = qcew[qcew["sector_idx"] == sector_idx]
        summary = summarize(sector_frame)
        by_sector.append(
            {
                "sector_idx": sector_idx,
                "sector_name": meta["name"],
                **summary,
            }
        )

    return {
        "overall_all_industries": summarize(qcew),
        "overall_cdp_sectors": summarize(qcew[cdp_mask]),
        "by_sector": by_sector,
    }


def compute_state_sector_detailed(industry_df: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    qcew = load_qcew_state_industry()
    available_codes = set(industry_df["naics_sig"].astype(str))
    coverage = qcew_coverage_metadata(qcew, available_codes)
    overall = coverage["overall_all_industries"]
    overall_cdp = coverage["overall_cdp_sectors"]
    log(
        "QCEW to national-industry employment match share "
        f"(all industries, exact/prefix): {overall['exact_match_share']:.4f} / {overall['prefix_match_share']:.4f}"
    )
    log(
        "QCEW to national-industry employment match share "
        f"(CDP sectors, exact/prefix): {overall_cdp['exact_match_share']:.4f} / {overall_cdp['prefix_match_share']:.4f}"
    )

    qcew["matched_naics_sig"] = qcew["naics_sig"].astype(str).map(lambda code: longest_available_prefix(code, available_codes))
    merged = qcew.merge(
        industry_df[["naics_sig", "sector_idx", "sector_name", "cognitive_intensity_raw"]].rename(
            columns={"naics_sig": "matched_naics_sig"}
        ),
        on="matched_naics_sig",
        how="left",
    )
    match_share = overall["prefix_match_share"]
    if not np.isfinite(match_share) or match_share < 0.40:
        raise RuntimeError(f"QCEW merge match share {match_share:.4f} is too low for the detailed state path.")
    if match_share < 0.75:
        log("Warning: QCEW match share is below 0.75; proceeding because detailed state-sector coverage is still testable.")

    merged = merged[merged["cognitive_intensity_raw"].notna()].copy()
    rows: list[dict] = []
    for state_fips, state_group in merged.groupby("state_fips", sort=True):
        state_abbr, state_idx = FIPS_TO_STATE[state_fips]
        for sector_idx, sector_meta in CDP_SECTOR_MAP.items():
            sector_group = state_group[state_group["sector_idx"] == sector_idx]
            if sector_group.empty:
                raise RuntimeError(f"Detailed state path is missing sector {sector_idx} for {state_abbr}")
            rows.append(
                {
                    "state_fips": state_fips,
                    "state_idx": state_idx,
                    "state_abbr": state_abbr,
                    "sector_idx": sector_idx,
                    "sector_name": sector_meta["name"],
                    "employment": float(sector_group["employment"].sum()),
                    "cognitive_intensity_raw": weighted_mean(
                        sector_group["cognitive_intensity_raw"], sector_group["employment"]
                    ),
                }
            )
    detailed = pd.DataFrame(rows).sort_values(["state_idx", "sector_idx"]).reset_index(drop=True)
    detailed.to_csv(INTERMEDIATE_DIR / "state_sector_cognitive_detailed.csv", index=False)
    log(f"Wrote detailed state-sector intensities to {INTERMEDIATE_DIR / 'state_sector_cognitive_detailed.csv'}")
    return detailed, {
        "state_path_mode": "detailed",
        "qcew_coverage": coverage,
    }


def locate_state_oews_workbook() -> Path:
    patterns = ["*23st*.zip", "*state*2023*.xlsx", "*oes*2023*state*.xlsx"]
    for pattern in patterns:
        matches = sorted(INPUT_DIR.rglob(pattern))
        if not matches:
            continue
        for match in matches:
            if match.suffix.lower() == ".xlsx":
                return match
            if match.suffix.lower() == ".zip":
                with zipfile.ZipFile(match) as archive:
                    workbook_members = [name for name in archive.namelist() if name.lower().endswith(".xlsx")]
                    if workbook_members:
                        member = sorted(workbook_members)[0]
                        dest = INPUT_DIR / Path(member).name
                        if not dest.exists():
                            with archive.open(member) as src, dest.open("wb") as dst:
                                dst.write(src.read())
                        return dest
    raise FileNotFoundError(
        "Detailed state path failed and no 2023 state OEWS workbook was found. "
        "Place `oesm23st.zip` or the extracted state workbook in input/."
    )


def infer_state_abbr(row: pd.Series) -> str | None:
    for column in ("PRIM_STATE", "AREA_TITLE", "AREA"):
        if column not in row.index:
            continue
        value = row[column]
        if value is None or (isinstance(value, float) and np.isnan(value)):
            continue
        text = str(value).strip()
        if not text:
            continue
        upper = text.upper()
        if upper in STATE_ABBR_TO_IDX:
            return upper
        if upper in STATE_NAME_TO_ABBR:
            return STATE_NAME_TO_ABBR[upper]
        code = canonicalize_code(text)
        if code and len(code) >= 2 and code[:2] in FIPS_TO_STATE:
            return FIPS_TO_STATE[code[:2]][0]
    return None


def compute_state_sector_fallback(
    occ_scores: pd.DataFrame,
    sector_df: pd.DataFrame,
    national_overall: float,
) -> tuple[pd.DataFrame, dict]:
    workbook_path = locate_state_oews_workbook()
    log(f"Using 2023 state OEWS fallback workbook {workbook_path}")
    raw = read_excel_table(workbook_path, STATE_OEWS_COLUMN_ALIASES)
    missing = {"OCC_CODE", "TOT_EMP"} - set(raw.columns)
    if missing:
        raise RuntimeError(f"State OEWS fallback workbook is missing required columns: {sorted(missing)}")

    df = raw.copy()
    if "O_GROUP" in df.columns:
        df["o_group"] = df["O_GROUP"].astype(str).str.strip().str.lower()
        df = df[df["o_group"] == "detailed"].copy()
    df["soc_code"] = df["OCC_CODE"].astype(str).str.strip()
    df = df[df["soc_code"].str.fullmatch(r"\d{2}-\d{4}")].copy()
    df["state_abbr"] = df.apply(infer_state_abbr, axis=1)
    df = df[df["state_abbr"].isin(STATE_ABBR_TO_IDX)].copy()
    df["tot_emp"] = to_numeric(df["TOT_EMP"])
    df = df[df["tot_emp"] > 0].copy()
    df = df.merge(occ_scores[["soc_code", "cognitive_score"]], on="soc_code", how="left")
    df = df[df["cognitive_score"].notna()].copy()
    state_rows: list[dict] = []
    for state_abbr, group in df.groupby("state_abbr", sort=True):
        state_rows.append(
            {
                "state_abbr": state_abbr,
                "state_idx": STATE_ABBR_TO_IDX[state_abbr],
                "state_cognitive_score": weighted_mean(group["cognitive_score"], group["tot_emp"]),
            }
        )
    state_df = pd.DataFrame(state_rows)
    if state_df["state_abbr"].nunique() != 50:
        raise RuntimeError("State OEWS fallback did not resolve to all 50 states.")

    rows: list[dict] = []
    for state in state_df.itertuples(index=False):
        for sector in sector_df.itertuples(index=False):
            rows.append(
                {
                    "state_idx": state.state_idx,
                    "state_abbr": state.state_abbr,
                    "sector_idx": sector.sector_idx,
                    "sector_name": sector.sector_name,
                    "employment": float("nan"),
                    "cognitive_intensity_raw": sector.cognitive_intensity_raw
                    * (state.state_cognitive_score / national_overall),
                }
            )
    fallback = pd.DataFrame(rows).sort_values(["state_idx", "sector_idx"]).reset_index(drop=True)
    fallback.to_csv(INTERMEDIATE_DIR / "state_sector_cognitive_fallback.csv", index=False)
    log(f"Wrote fallback state-sector intensities to {INTERMEDIATE_DIR / 'state_sector_cognitive_fallback.csv'}")
    return fallback, {
        "state_path_mode": "fallback",
        "qcew_coverage": None,
    }


def build_output_metadata(final_df: pd.DataFrame, provenance: list[dict], state_path_details: dict) -> dict:
    source_files = {
        row.get("dataset"): Path(row["local_path"]).name
        for row in provenance
        if row.get("dataset") and row.get("dataset") != "summary" and row.get("local_path")
    }
    metadata = {
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "source_files": source_files,
        "download_provenance": provenance,
        "state_path_mode": state_path_details.get("state_path_mode"),
        "qcew_coverage": state_path_details.get("qcew_coverage"),
        "detailed_path_error": state_path_details.get("detailed_path_error"),
        "matrix_summary": {
            "row_count": int(len(final_df)),
            "state_count": int(final_df["state_idx"].nunique()),
            "sector_count": int(final_df["sector_idx"].nunique()),
            "cognitive_intensity_min": float(final_df["cognitive_intensity"].min()),
            "cognitive_intensity_max": float(final_df["cognitive_intensity"].max()),
            "cognitive_intensity_raw_min": float(final_df["cognitive_intensity_raw"].min()),
            "cognitive_intensity_raw_max": float(final_df["cognitive_intensity_raw"].max()),
        },
    }
    return metadata


def finalize_outputs(raw_state_sector: pd.DataFrame, provenance: list[dict], state_path_details: dict) -> None:
    final_df = raw_state_sector.copy().sort_values(["state_idx", "sector_idx"]).reset_index(drop=True)
    if len(final_df) != 200:
        raise RuntimeError(f"Expected 200 state-sector rows, found {len(final_df)}")
    min_raw = final_df["cognitive_intensity_raw"].min()
    max_raw = final_df["cognitive_intensity_raw"].max()
    if pd.isna(min_raw) or pd.isna(max_raw) or max_raw <= min_raw:
        raise RuntimeError("Final normalization failed because raw intensities are degenerate.")
    final_df["cognitive_intensity"] = (final_df["cognitive_intensity_raw"] - min_raw) / (max_raw - min_raw)
    final_df = final_df[
        ["state_idx", "state_abbr", "sector_idx", "sector_name", "cognitive_intensity", "cognitive_intensity_raw"]
    ]
    final_df.to_csv(OUTPUT_DIR / "cognitive_intensity_matrix.csv", index=False)

    wide = (
        final_df.pivot(index=["state_idx", "state_abbr"], columns="sector_name", values="cognitive_intensity")
        .reset_index()
        .sort_values("state_idx")
    )
    wide.to_csv(OUTPUT_DIR / "cognitive_intensity_wide.csv", index=False)
    metadata = build_output_metadata(final_df, provenance, state_path_details)
    METADATA_PATH.write_text(json.dumps(metadata, indent=2))
    log(f"Wrote final outputs to {OUTPUT_DIR / 'cognitive_intensity_matrix.csv'} and {OUTPUT_DIR / 'cognitive_intensity_wide.csv'}")
    log(f"Wrote metadata to {METADATA_PATH}")


def main() -> None:
    for path in (INTERMEDIATE_DIR, OUTPUT_DIR):
        path.mkdir(parents=True, exist_ok=True)

    provenance = load_download_provenance()
    if provenance:
        summary_rows = [row for row in provenance if row.get("dataset") == "summary"]
        if summary_rows and summary_rows[-1].get("status") != "success":
            raise RuntimeError("Download provenance indicates the download step failed. Run `make download` first.")

    occ_scores = collapse_onet_scores()
    oews = load_oews_national_matrix()
    merged, _ = merge_oews_with_cognitive_scores(oews, occ_scores)
    industry_df, sector_df, national_overall = compute_national_outputs(merged)

    try:
        raw_state_sector, state_path_details = compute_state_sector_detailed(industry_df)
    except Exception as exc:
        log(f"Detailed state path failed: {exc}")
        raw_state_sector, state_path_details = compute_state_sector_fallback(occ_scores, sector_df, national_overall)
        state_path_details["detailed_path_error"] = str(exc)

    finalize_outputs(raw_state_sector, provenance, state_path_details)


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        log(str(exc))
        sys.exit(1)
